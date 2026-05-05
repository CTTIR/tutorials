#!/usr/bin/env python3
"""Distribute VG Wort Zählmarken into Quarto .qmd pages.

Reads a shared xlsx inventory of unused markers, finds eligible .qmd pages
(>= --min-chars of prose), and embeds one marker per page atomically while
flipping the corresponding xlsx row to Used=True with the page URL.

Run --dry-run first; --apply requires a clean git tree.
"""
from __future__ import annotations

import argparse
import fnmatch
import os
import re
import shutil
import subprocess
import sys
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Optional

try:
    import openpyxl
except ImportError:
    sys.stderr.write("ERROR: openpyxl is required. Install with: pip install openpyxl\n")
    sys.exit(2)

try:
    import yaml
except ImportError:
    yaml = None  # type: ignore


# --------------------------------------------------------------------------- #
# Configuration
# --------------------------------------------------------------------------- #

DEFAULT_EXCLUDES = [
    "docs/**",
    "_freeze/**",
    ".quarto/**",
    "_extensions/**",
    "_inc/**",
    "assets/**",
]

VGWORT_HOST_RE = re.compile(r"vg\d{2}\.met\.vgwort\.de/na/([0-9a-f]{32})", re.IGNORECASE)
VGWORT_DIV_RE = re.compile(r":::\s*\{\.vgwort-pixel\}")
VGWORT_COMMENT_RE = re.compile(r"VG Wort Zählmarke", re.IGNORECASE)

XLSX_RETRY_ATTEMPTS = 3
XLSX_RETRY_BACKOFF_S = 2.0


# --------------------------------------------------------------------------- #
# Data classes
# --------------------------------------------------------------------------- #

@dataclass
class MarkerRow:
    row_index: int  # 1-based xlsx row (incl. header)
    public_id: str
    private_id: str
    html_code: str
    used: bool
    url: Optional[str]


@dataclass
class QmdInfo:
    path: Path
    rel_path: str
    target_url: str
    has_marker: bool
    existing_public_id: Optional[str]
    prose_chars: int
    is_draft: bool
    is_partial: bool  # leading underscore => Quarto include partial


# --------------------------------------------------------------------------- #
# Prose character counting
# --------------------------------------------------------------------------- #

def count_prose_chars(qmd_path: Path) -> int:
    text = qmd_path.read_text(encoding="utf-8")
    text = re.sub(r"\A---\n.*?\n---\n", "", text, count=1, flags=re.DOTALL)
    text = re.sub(r"```.*?\n.*?```", "", text, flags=re.DOTALL)
    text = re.sub(r"<!--.*?-->", "", text, flags=re.DOTALL)
    text = re.sub(r"\$\$.*?\$\$", "", text, flags=re.DOTALL)
    text = re.sub(r"\$[^$\n]+\$", "", text)
    text = re.sub(r"\{\{<.*?>\}\}", "", text, flags=re.DOTALL)
    text = re.sub(r"`[^`\n]+`", "", text)
    text = re.sub(r"<[^>]+>", "", text)
    text = re.sub(r"!\[([^\]]*)\]\([^)]*\)", r"\1", text)
    text = re.sub(r"\[([^\]]+)\]\([^)]*\)", r"\1", text)
    text = re.sub(r"^:::.*$", "", text, flags=re.MULTILINE)
    text = re.sub(r"\s+", " ", text).strip()
    return len(text)


# --------------------------------------------------------------------------- #
# Frontmatter / draft detection
# --------------------------------------------------------------------------- #

def is_draft(qmd_path: Path) -> bool:
    text = qmd_path.read_text(encoding="utf-8")
    m = re.match(r"\A---\n(.*?)\n---\n", text, flags=re.DOTALL)
    if not m:
        return False
    fm = m.group(1)
    if yaml is not None:
        try:
            data = yaml.safe_load(fm)
            if isinstance(data, dict) and data.get("draft") is True:
                return True
        except Exception:
            pass
    return bool(re.search(r"^\s*draft\s*:\s*true\s*$", fm, flags=re.MULTILINE | re.IGNORECASE))


# --------------------------------------------------------------------------- #
# Existing marker detection
# --------------------------------------------------------------------------- #

def detect_existing_marker(qmd_path: Path) -> tuple[bool, Optional[str]]:
    text = qmd_path.read_text(encoding="utf-8")
    m = VGWORT_HOST_RE.search(text)
    if m:
        return True, m.group(1).lower()
    if VGWORT_DIV_RE.search(text) or VGWORT_COMMENT_RE.search(text):
        return True, None
    return False, None


# --------------------------------------------------------------------------- #
# Quarto site-url & URL building
# --------------------------------------------------------------------------- #

def read_site_url(root: Path) -> str:
    qy = root / "_quarto.yml"
    if not qy.exists():
        sys.stderr.write(f"ERROR: {qy} not found — not a Quarto project root?\n")
        sys.exit(2)
    text = qy.read_text(encoding="utf-8")
    if yaml is not None:
        try:
            data = yaml.safe_load(text) or {}
            site_url = (data.get("website") or {}).get("site-url")
            if site_url:
                return site_url.rstrip("/") + "/"
        except Exception:
            pass
    m = re.search(r"site-url\s*:\s*[\"']?([^\"'\n]+)[\"']?", text)
    if not m:
        sys.stderr.write("ERROR: could not read website.site-url from _quarto.yml\n")
        sys.exit(2)
    return m.group(1).strip().rstrip("/") + "/"


def build_target_url(site_url: str, rel_path: str) -> str:
    rel = rel_path.replace("\\", "/").lstrip("./").lstrip("/")
    if rel.endswith(".qmd"):
        rel = rel[:-4] + ".html"
    return site_url + rel


# --------------------------------------------------------------------------- #
# xlsx I/O
# --------------------------------------------------------------------------- #

EXPECTED_HEADERS = [
    "Öffentlicher Identifikationscode",
    "Privater Identifikationscode",
    "HTML_Code",
    "Used",
    "URL",
]


def _open_workbook(xlsx_path: Path):
    last_err = None
    for attempt in range(1, XLSX_RETRY_ATTEMPTS + 1):
        try:
            return openpyxl.load_workbook(xlsx_path)
        except (PermissionError, OSError) as e:
            last_err = e
            if attempt < XLSX_RETRY_ATTEMPTS:
                time.sleep(XLSX_RETRY_BACKOFF_S)
    raise RuntimeError(f"Could not open {xlsx_path} after {XLSX_RETRY_ATTEMPTS} attempts: {last_err}")


def load_inventory(xlsx_path: Path):
    wb = _open_workbook(xlsx_path)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
    headers = [c.value for c in ws[1]]
    for h in EXPECTED_HEADERS:
        if h not in headers:
            raise RuntimeError(f"xlsx missing expected column: {h!r}. Found: {headers}")
    col = {h: headers.index(h) + 1 for h in EXPECTED_HEADERS}

    rows: dict[str, MarkerRow] = {}
    for r in range(2, ws.max_row + 1):
        pub = ws.cell(r, col["Öffentlicher Identifikationscode"]).value
        if not pub:
            continue
        pub_norm = str(pub).strip().lower()
        used_val = ws.cell(r, col["Used"]).value
        used = bool(used_val) and str(used_val).strip().lower() not in ("false", "0", "")
        rows[pub_norm] = MarkerRow(
            row_index=r,
            public_id=pub_norm,
            private_id=str(ws.cell(r, col["Privater Identifikationscode"]).value or ""),
            html_code=str(ws.cell(r, col["HTML_Code"]).value or ""),
            used=used,
            url=ws.cell(r, col["URL"]).value,
        )
    return wb, ws, col, rows


def atomic_save_workbook(wb, xlsx_path: Path):
    tmp = xlsx_path.with_suffix(xlsx_path.suffix + ".tmp")
    last_err = None
    for attempt in range(1, XLSX_RETRY_ATTEMPTS + 1):
        try:
            wb.save(tmp)
            try:
                fd = os.open(tmp, os.O_RDWR)
                try:
                    os.fsync(fd)
                finally:
                    os.close(fd)
            except OSError:
                pass
            os.replace(tmp, xlsx_path)
            return
        except (PermissionError, OSError) as e:
            last_err = e
            if attempt < XLSX_RETRY_ATTEMPTS:
                time.sleep(XLSX_RETRY_BACKOFF_S)
    if tmp.exists():
        try:
            tmp.unlink()
        except OSError:
            pass
    raise RuntimeError(f"Could not save xlsx after {XLSX_RETRY_ATTEMPTS} attempts: {last_err}")


# --------------------------------------------------------------------------- #
# Snippet rendering
# --------------------------------------------------------------------------- #

def render_snippet(html_code: str, public_id: str) -> str:
    img_block = html_code.replace("\\n", "\n").strip()
    if not img_block:
        img_block = (
            f'<img src="https://vg09.met.vgwort.de/na/{public_id}" '
            f'width="1" height="1" alt="" loading="eager" fetchpriority="high" '
            f'decoding="async" style="position:absolute;visibility:hidden;" />'
        )
    return (
        "\n"
        "::: {.vgwort-pixel}\n"
        "```{=html}\n"
        f"<!-- VG Wort Zählmarke — public ID: {public_id} -->\n"
        f"{img_block}\n"
        "```\n"
        ":::\n"
    )


def append_snippet(qmd_path: Path, snippet: str) -> None:
    text = qmd_path.read_text(encoding="utf-8")
    if not text.endswith("\n"):
        text += "\n"
    text += snippet
    qmd_path.write_text(text, encoding="utf-8")


# --------------------------------------------------------------------------- #
# Discovery
# --------------------------------------------------------------------------- #

def matches_any(rel: str, patterns: list[str]) -> bool:
    rel_norm = rel.replace("\\", "/")
    for p in patterns:
        if fnmatch.fnmatch(rel_norm, p) or fnmatch.fnmatch(rel_norm, p.rstrip("/*") + "/*"):
            return True
    return False


def discover_qmds(root: Path, excludes: list[str], onlys: list[str]) -> list[Path]:
    out = []
    for p in root.rglob("*.qmd"):
        rel = p.relative_to(root).as_posix()
        if matches_any(rel, excludes):
            continue
        if onlys and not matches_any(rel, onlys):
            continue
        out.append(p)
    return sorted(out)


# --------------------------------------------------------------------------- #
# Git safety
# --------------------------------------------------------------------------- #

def assert_clean_git(root: Path) -> None:
    try:
        out = subprocess.check_output(
            ["git", "status", "--porcelain"], cwd=root, text=True, stderr=subprocess.STDOUT
        )
    except (subprocess.CalledProcessError, FileNotFoundError) as e:
        sys.stderr.write(f"ERROR: cannot run `git status`: {e}\n")
        sys.exit(2)
    if out.strip():
        sys.stderr.write("ERROR: working tree not clean. Commit/stash before --apply:\n")
        sys.stderr.write(out)
        sys.exit(2)


def check_dropbox_conflict(xlsx_path: Path) -> None:
    parent = xlsx_path.parent
    stem = xlsx_path.stem
    for sib in parent.glob(f"{stem}*CONFLICTED COPY*{xlsx_path.suffix}"):
        sys.stderr.write(f"ERROR: Dropbox conflict file detected: {sib.name}\n")
        sys.stderr.write("Reconcile manually before re-running.\n")
        sys.exit(2)


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #

def parse_args() -> argparse.Namespace:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--xlsx", required=True, type=Path)
    mode = ap.add_mutually_exclusive_group()
    mode.add_argument("--dry-run", action="store_true", default=True)
    mode.add_argument("--apply", action="store_true")
    ap.add_argument("--root", type=Path, default=Path("."))
    ap.add_argument("--min-chars", type=int, default=1800)
    ap.add_argument("--exclude", action="append", default=None)
    ap.add_argument("--only", action="append", default=None)
    args = ap.parse_args()
    if args.apply:
        args.dry_run = False
    return args


def main() -> int:
    args = parse_args()
    root = args.root.resolve()
    excludes = list(DEFAULT_EXCLUDES) + list(args.exclude or [])
    onlys = list(args.only or [])

    site_url = read_site_url(root)

    if args.apply:
        assert_clean_git(root)
    check_dropbox_conflict(args.xlsx)

    if not args.xlsx.exists():
        sys.stderr.write(f"ERROR: xlsx not found: {args.xlsx}\n")
        return 2

    wb, ws, col, rows_by_id = load_inventory(args.xlsx)
    initial_available = sum(1 for mr in rows_by_id.values() if not mr.used)

    print(f"Inventory: {len(rows_by_id)} markers loaded, {initial_available} available")
    print(f"Site URL : {site_url}")
    print(f"Root     : {root}")
    print(f"Mode     : {'APPLY' if args.apply else 'DRY RUN'}")
    print()

    qmds = discover_qmds(root, excludes, onlys)

    infos: list[QmdInfo] = []
    for p in qmds:
        rel = p.relative_to(root).as_posix()
        is_partial = p.name.startswith("_")
        draft = is_draft(p)
        has_m, pub_id = detect_existing_marker(p)
        chars = count_prose_chars(p)
        infos.append(QmdInfo(
            path=p, rel_path=rel,
            target_url=build_target_url(site_url, rel),
            has_marker=has_m, existing_public_id=pub_id,
            prose_chars=chars, is_draft=draft, is_partial=is_partial,
        ))

    print("PHASE 1 — discovery")
    width = max((len(i.rel_path) for i in infos), default=20)
    for i in infos:
        if i.is_partial:
            tag = "PARTIAL (skip)"
        elif i.is_draft:
            tag = "DRAFT (skip)"
        elif i.has_marker:
            tag = "RECONCILE"
        elif i.prose_chars >= args.min_chars:
            tag = "ELIGIBLE"
        else:
            tag = "TOO SHORT"
        print(f"  {i.rel_path.ljust(width)}  marked={'yes' if i.has_marker else 'no '}  chars={i.prose_chars:6d}  {tag}")
    print()

    # Phase 2: reconciliation
    print("PHASE 2 — reconciliation")
    reconcile_ok = 0
    reconcile_anomalies = 0
    xlsx_dirty = False
    for i in infos:
        if i.is_partial or i.is_draft or not i.has_marker:
            continue
        if not i.existing_public_id:
            print(f"  WARN  {i.rel_path}: marker present but public ID could not be extracted")
            reconcile_anomalies += 1
            continue
        mr = rows_by_id.get(i.existing_public_id)
        if mr is None:
            print(f"  ERROR {i.rel_path}: public ID {i.existing_public_id} not in xlsx (foreign marker)")
            reconcile_anomalies += 1
            continue
        if mr.used:
            if mr.url and mr.url != i.target_url:
                print(f"  WARN  {i.rel_path}: xlsx URL mismatch (xlsx={mr.url} vs page={i.target_url})")
                reconcile_anomalies += 1
            else:
                reconcile_ok += 1
        else:
            print(f"  FIX   {i.rel_path}: xlsx Used=False but marker is in qmd; setting Used=True")
            if args.apply:
                ws.cell(mr.row_index, col["Used"]).value = True
                ws.cell(mr.row_index, col["URL"]).value = i.target_url
                mr.used = True
                mr.url = i.target_url
                xlsx_dirty = True
            reconcile_ok += 1
    print()

    # Phase 3: allocation — rebuild `available` AFTER reconciliation so any
    # row that Phase 2 just flipped to Used=True is excluded from allocation.
    available = sorted(
        (mr for mr in rows_by_id.values() if not mr.used), key=lambda x: x.row_index
    )

    print(f"PHASE 3 — allocation [{'APPLY' if args.apply else 'DRY RUN'}]")
    eligible = [
        i for i in infos
        if not i.is_partial and not i.is_draft
        and not i.has_marker and i.prose_chars >= args.min_chars
    ]

    if len(eligible) > len(available):
        print(f"  HALT: {len(eligible)} eligible pages but only {len(available)} markers available")
        print("  Refusing to partially allocate. Aborting before any write.")
        return 3

    allocated = 0
    backup_made = False
    if args.apply and eligible and not backup_made:
        ts = datetime.now().strftime("%Y%m%d-%H%M%S")
        backup = args.xlsx.with_name(f"{args.xlsx.name}.bak.{ts}")
        shutil.copy2(args.xlsx, backup)
        print(f"  BACKUP {backup}")
        backup_made = True

    avail_iter = iter(available)
    for i in eligible:
        try:
            mr = next(avail_iter)
        except StopIteration:
            print(f"  HALT: ran out of markers at {i.rel_path}")
            return 3
        snippet = render_snippet(mr.html_code, mr.public_id)
        if args.apply:
            append_snippet(i.path, snippet)
            ws.cell(mr.row_index, col["Used"]).value = True
            ws.cell(mr.row_index, col["URL"]).value = i.target_url
            mr.used = True
            mr.url = i.target_url
            atomic_save_workbook(wb, args.xlsx)
            print(f"  OK    {i.rel_path}  <-  {mr.public_id}")
        else:
            print(f"  WOULD WRITE  {i.rel_path}  <-  {mr.public_id}")
        allocated += 1

    if args.apply and xlsx_dirty and allocated == 0:
        # Phase 2 fixes only; persist them.
        atomic_save_workbook(wb, args.xlsx)
    print()

    # Phase 4: too short
    print("PHASE 4 — too short (informational)")
    too_short = [
        i for i in infos
        if not i.is_partial and not i.is_draft
        and not i.has_marker and i.prose_chars < args.min_chars
    ]
    for i in too_short:
        print(f"  SKIP (too short, {i.prose_chars} chars): {i.rel_path}")
    print()

    # Phase 5: summary
    remaining = sum(1 for mr in rows_by_id.values() if not mr.used)
    print("SUMMARY")
    print(f"  Discovered:           {len(infos)} qmd files")
    print(f"  Already marked:       {sum(1 for i in infos if i.has_marker and not i.is_partial and not i.is_draft):>4d}  "
          f"({reconcile_ok} reconciled OK, {reconcile_anomalies} anomalies)")
    print(f"  Newly allocated:      {allocated:>4d}")
    print(f"  Too short:            {len(too_short):>4d}")
    print(f"  Drafts/partials:      {sum(1 for i in infos if i.is_partial or i.is_draft):>4d}")
    print(f"  Available remaining:  {remaining:>4d}")
    if args.dry_run:
        print("  DRY RUN — re-run with --apply to write changes")
    else:
        print("  DONE — review with `git diff` and commit")

    if reconcile_anomalies > 0:
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
